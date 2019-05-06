# -*- coding: utf-8 -*-
"""
Created on Mon Apr 29 19:56:53 2019

@author: 91727
"""

import re
from openpyxl import load_workbook

wb = load_workbook("dataset.xlsx")
sheet = wb.worksheets[0]

#print(sheet.max_row)

for i in range(2,sheet.max_row+1):
    install = str(sheet.cell(row= i, column=5).value)
    if(install[-1]=='M'):
        sheet.cell(row=i,column=5).value = install[:-1]

wb.save("dataset.xlsx")

for i in range(2,sheet.max_row+1):
    size = str(sheet.cell(row= i, column=4).value)
    value=0
    if(size[0]>='0' and size[0]<='9'):
        length = len(size)
        if(size[length-1]=='k'):
            value = float(str(size[:-1]))
            value = value*1024
            #print(value)
        elif(size[length-1]=='M'):
            value = float(str(size[:-1]))
            value = value*1024*1024
            #print(value)
            
    sheet.cell(row=i,column=4).value = value
wb.save("dataset.xlsx")

for i in range(2,sheet.max_row+1):
    price = str(sheet.cell(row= i, column=6).value)
    if(price[0]=='$'):
        sheet.cell(row=i,column=6).value = price[1:]
        print(price)
wb.save("dataset.xlsx")



"""
for i in range(2,sheet.max_row+1):
    size = str(sheet.cell(row= i, column=6).value)
    value=''
    if(size[0]>='0' and size[0]<='9'):
        length = len(size)
        if(size[length-1]=='+'):
            value=size[:-1]
            print(value)
        else:
            print(i, size)        
    sheet.cell(row=i,column=6).value = value
wb.save("googleplaystore.xlsx")
"""
"""
for i in range(2,sheet.max_row+1):
    size = str(sheet.cell(row= i, column=5).value)
    value=0
    if(size[0]>='0' and size[0]<='9'):
        length = len(size)
        if(size[length-1]=='k'):
            value = float(str(size[:-1]))
            value = value*1024
            print(value)
        elif(size[length-1]=='M'):
            value = float(str(size[:-1]))
            value = value*1024*1024
            print(value)
            
    sheet.cell(row=i,column=5).value = value
wb.save("googleplaystore.xlsx")
"""

